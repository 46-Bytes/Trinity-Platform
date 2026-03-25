"""
Strategy Workbook Excel exporter.

Loads the Strategy Workbook Template and maps extracted data to the correct cells,
preserving all formatting, dropdowns, and data validation.
"""
from __future__ import annotations

from typing import Dict, Any, List, Optional
from pathlib import Path
import io
import logging
import copy

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class StrategyWorkbookExporter:
    """
    Export extracted strategy data to Excel workbook.
    Loads template and maps data while preserving formatting.
    """
    
    def __init__(self, template_path: Optional[Path] = None):
        """
        Initialize exporter.
        
        Args:
            template_path: Path to template file. If None, uses default location.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )
        
        if template_path is None:
            # Default template location
            base_dir = Path(__file__).resolve().parents[2]  # Go up to backend/
            template_path = base_dir / "files" / "templates" / "strategy-workbook" / "Strategy Workbook Template.xlsx"
        
        self.template_path = Path(template_path)
        
        if not self.template_path.exists():
            raise FileNotFoundError(
                f"Template file not found at {self.template_path}. "
                f"Please place 'Strategy Workbook Template.xlsx' in the templates directory."
            )
    
    @staticmethod
    def _safe_set_value(ws, row: int, column: int, value):
        """
        Safely set a cell value, handling merged cells.
        If the target cell is part of a merged range, writes to the
        top-left (primary) cell of that range instead.
        """
        cell = ws.cell(row=row, column=column)
        try:
            cell.value = value
        except AttributeError:
            # Cell is a MergedCell (read-only). Find the merge range and write to its top-left cell.
            for merge_range in ws.merged_cells.ranges:
                if merge_range.min_row <= row <= merge_range.max_row and merge_range.min_col <= column <= merge_range.max_col:
                    ws.cell(row=merge_range.min_row, column=merge_range.min_col).value = value
                    return

    def generate_workbook(
        self,
        extracted_data: Dict[str, Any],
        output_path: Optional[Path] = None
    ) -> bytes:
        """
        Generate workbook from extracted data.
        
        Args:
            extracted_data: Extracted data dictionary
            output_path: Optional path to save workbook. If None, returns bytes.
            
        Returns:
            Workbook file bytes
        """
        # Load template
        wb = load_workbook(self.template_path)
        
        # Get the main worksheet (assume first sheet or find by name)
        # Template should have a main sheet - we'll use the active sheet
        ws = wb.active
        
        logger.info(f"Loaded template with {len(wb.sheetnames)} sheets")
        
        # Map data to workbook
        self._map_visioning(ws, extracted_data.get("visioning", {}))
        self._map_business_model(ws, extracted_data.get("business_model", {}))
        self._map_market_segmentation(ws, extracted_data.get("market_segmentation", []))
        self._map_porters_5_forces(ws, extracted_data.get("porters_5_forces", []))
        self._map_pestel(ws, extracted_data.get("pestel", []))
        self._map_swot(ws, extracted_data.get("swot", {}))
        self._map_customer_analysis(ws, extracted_data.get("customer_analysis", []))
        self._map_product_analysis(ws, extracted_data.get("product_analysis", []))
        self._map_competitor_analysis(ws, extracted_data.get("competitor_analysis", []))
        self._map_growth_opportunities(ws, extracted_data.get("growth_opportunities", []))
        self._map_financial_targets(ws, extracted_data.get("financial_targets", {}))
        self._map_risks(ws, extracted_data.get("risks", {}))
        self._map_strategic_priorities(ws, extracted_data.get("strategic_priorities", []))
        self._map_key_actions(ws, extracted_data.get("key_actions", []))
        
        # Save to bytes or file
        if output_path:
            wb.save(output_path)
            with open(output_path, "rb") as f:
                return f.read()
        else:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
    
    def _find_section_start(self, ws, section_name: str) -> Optional[int]:
        """
        Find the row number where a section starts.
        
        Args:
            ws: Worksheet
            section_name: Name of section to find (e.g., "VISIONING")
            
        Returns:
            Row number or None if not found
        """
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=500, values_only=False), start=1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if section_name.upper() in cell.value.upper():
                        return row_idx
        return None
    
    def _find_next_section_start(self, ws, after_row: int) -> Optional[int]:
        """
        Find the row number of the next section header after a given row.
        
        Args:
            ws: Worksheet
            after_row: Row to search after
            
        Returns:
            Row number of next section or None
        """
        for row_idx, row in enumerate(ws.iter_rows(min_row=after_row+1, max_row=after_row+100, values_only=False), start=after_row+1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Check if it looks like a section header (all caps, short text)
                    text = str(cell.value).strip()
                    if text.isupper() and len(text) < 50 and len(text) > 3:
                        # Check if cell is bold (likely a header)
                        if cell.font and cell.font.bold:
                            return row_idx
        return None
    
    def _find_data_start_row(self, ws, section_start_row: int) -> int:
        """
        Find the first data row after a section header.
        Usually the row after the header or after column headers.
        
        Args:
            ws: Worksheet
            section_start_row: Row where section starts
            
        Returns:
            Row number where data should start
        """
        # Look for table headers or first empty row after section header
        for row_idx in range(section_start_row + 1, section_start_row + 10):
            row = ws[row_idx]
            # Check if this row has column headers (non-empty cells in first few columns)
            has_headers = any(cell.value for cell in row[:5])
            if has_headers:
                return row_idx + 1  # Data starts after headers
        return section_start_row + 2  # Default: skip header row
    
    def _insert_row_with_formatting(self, ws, row_idx: int, source_row_idx: int):
        """
        Insert a new row and copy formatting from source row.
        
        Args:
            ws: Worksheet
            row_idx: Row index to insert at
            source_row_idx: Row index to copy formatting from
        """
        ws.insert_rows(row_idx)
        
        # Copy formatting from source row
        source_row = ws[source_row_idx]
        new_row = ws[row_idx]
        
        for col_idx, (source_cell, new_cell) in enumerate(zip(source_row, new_row), start=1):
            # Copy style attributes individually (StyleProxy objects can't be directly assigned)
            if source_cell.has_style:
                # Copy font - create new Font object with same attributes
                if source_cell.font:
                    try:
                        font = source_cell.font
                        new_cell.font = Font(
                            name=font.name,
                            size=font.size,
                            bold=font.bold,
                            italic=font.italic,
                            vertAlign=font.vertAlign,
                            underline=font.underline,
                            strike=font.strike,
                            color=font.color
                        )
                    except Exception:
                        pass  # Skip if font copying fails
                
                # Copy fill - create new PatternFill with same attributes
                if source_cell.fill:
                    try:
                        fill = source_cell.fill
                        if hasattr(fill, 'patternType') and fill.patternType:
                            new_cell.fill = PatternFill(
                                fill_type=fill.patternType,
                                start_color=fill.start_color,
                                end_color=fill.end_color
                            )
                    except Exception:
                        pass  # Skip if fill copying fails
                
                # Copy border - create new Border with same attributes
                if source_cell.border:
                    try:
                        border = source_cell.border
                        new_cell.border = Border(
                            left=border.left,
                            right=border.right,
                            top=border.top,
                            bottom=border.bottom
                        )
                    except Exception:
                        pass  # Skip if border copying fails
                
                # Copy alignment - create new Alignment with same attributes
                if source_cell.alignment:
                    try:
                        align = source_cell.alignment
                        new_cell.alignment = Alignment(
                            horizontal=align.horizontal,
                            vertical=align.vertical,
                            text_rotation=align.text_rotation,
                            wrap_text=align.wrap_text,
                            shrink_to_fit=align.shrink_to_fit,
                            indent=align.indent
                        )
                    except Exception:
                        pass  # Skip if alignment copying fails
                
                # Copy number format (string, so safe to assign directly)
                if source_cell.number_format:
                    new_cell.number_format = source_cell.number_format
                
                # Copy protection - create new Protection with same attributes
                if source_cell.protection:
                    try:
                        prot = source_cell.protection
                        from openpyxl.styles import Protection
                        new_cell.protection = Protection(
                            locked=prot.locked,
                            hidden=prot.hidden
                        )
                    except Exception:
                        pass  # Skip if protection copying fails
            
            # Copy data validation (if any)
            # Note: openpyxl doesn't easily copy validation, so we'll handle it per section
    
    def _map_visioning(self, ws, visioning_data: Dict[str, Any]):
        """Map visioning data to Column B text fields."""
        section_row = self._find_section_start(ws, "VISIONING")
        if not section_row:
            logger.warning("VISIONING section not found in template")
            return
        
        # Map each field to its corresponding row in Column B
        # This assumes the template has labels in Column A and fields in Column B
        field_mapping = {
            "business_description": "Describe the business that you are in",
            "business_size_goal": "How big do you want the business to be",
            "markets_serviced": "What markets will it service",
            "customer_groups": "What industries or groups of customers",
            "geographic_spread": "What is the geographic spread",
            "facilities_count": "How many factories / offices will there be",
            "facilities_locations": "Where will they be",
            "future_comparison": "What will they look like compared with now",
            "customer_description": "How will your customers describe the business",
            "competitor_description": "How will your competitors describe the business",
            "employee_description": "How will your employees describe the business",
            "achievements": "What achievements will the business have made",
            "exit_intent": "Exit / What is your planned involvement with the business"
        }
        
        # Search for each field label and map value
        for field_key, label_text in field_mapping.items():
            value = visioning_data.get(field_key)
            if value:
                # Find row with this label
                for row_idx in range(section_row, section_row + 20):
                    cell_a = ws.cell(row=row_idx, column=1)
                    if cell_a.value and label_text.lower() in str(cell_a.value).lower():
                        self._safe_set_value(ws, row_idx, 2, str(value))
                        break
    
    def _map_business_model(self, ws, business_model_data: Dict[str, Any]):
        """Map business model data to Column C table rows."""
        section_row = self._find_section_start(ws, "BUSINESS MODEL")
        if not section_row:
            logger.warning("BUSINESS MODEL section not found")
            return
        
        data_start = self._find_data_start_row(ws, section_row)
        
        # Map to Column C for each category
        mappings = {
            "revenue_streams": business_model_data.get("revenue_streams", []),
            "key_products_services": business_model_data.get("key_products_services", []),
            "customer_segments": business_model_data.get("customer_segments", []),
            "delivery_model": [business_model_data.get("delivery_model")] if business_model_data.get("delivery_model") else [],
            "key_partners": business_model_data.get("key_partners", []),
            "key_cost_drivers": business_model_data.get("key_cost_drivers", [])
        }
        
        current_row = data_start
        for category, items in mappings.items():
            if items:
                # Find the row for this category or insert
                for row_idx in range(data_start, data_start + 20):
                    cell = ws.cell(row=row_idx, column=1)
                    if cell.value and category.replace("_", " ").lower() in str(cell.value).lower():
                        # Write items to Column C, using line breaks for multiple
                        self._safe_set_value(ws, row_idx, 3, "\n".join(str(item) for item in items if item))
                        break
    
    def _map_market_segmentation(self, ws, segments: List[Dict[str, Any]]):
        """Map market segmentation data - insert rows for each segment."""
        section_row = self._find_section_start(ws, "MARKET SEGMENTATION")
        if not section_row:
            logger.warning("MARKET SEGMENTATION section not found")
            return
        
        data_start = self._find_data_start_row(ws, section_row)
        
        # Find last row in section
        next_section = self._find_next_section_start(ws, section_row)
        last_row = next_section - 1 if next_section else data_start + 50
        
        # Find the last data row (first empty row in data area)
        insert_row = data_start
        for row_idx in range(data_start, last_row):
            if not any(ws.cell(row=row_idx, column=col).value for col in range(1, 8)):
                insert_row = row_idx
                break
            insert_row = row_idx + 1
        
        # Insert rows for each segment
        for segment in segments:
            if insert_row > last_row:
                # Need to insert before next section
                self._insert_row_with_formatting(ws, insert_row, insert_row - 1)
            
            self._safe_set_value(ws, insert_row, 1, segment.get("market_product_group"))
            self._safe_set_value(ws, insert_row, 2, segment.get("customer_needs"))
            self._safe_set_value(ws, insert_row, 3, segment.get("solution_sought"))
            self._safe_set_value(ws, insert_row, 4, segment.get("share_of_revenue_percent"))
            self._safe_set_value(ws, insert_row, 5, segment.get("growth_rating"))
            self._safe_set_value(ws, insert_row, 6, segment.get("profitability_rating"))
            self._safe_set_value(ws, insert_row, 7, segment.get("market_position"))
            
            insert_row += 1
    
    def _map_porters_5_forces(self, ws, forces: List[Dict[str, Any]]):
        """Map Porter's 5 Forces - insert rows for each observation."""
        section_row = self._find_section_start(ws, "PORTERS 5 FORCES")
        if not section_row:
            logger.warning("PORTERS 5 FORCES section not found")
            return
        
        data_start = self._find_data_start_row(ws, section_row)
        next_section = self._find_next_section_start(ws, section_row)
        last_row = next_section - 1 if next_section else data_start + 50
        
        insert_row = data_start
        for row_idx in range(data_start, last_row):
            if not any(ws.cell(row=row_idx, column=col).value for col in range(1, 5)):
                insert_row = row_idx
                break
            insert_row = row_idx + 1
        
        for force in forces:
            if insert_row > last_row:
                self._insert_row_with_formatting(ws, insert_row, insert_row - 1)
            
            self._safe_set_value(ws, insert_row, 1, force.get("force"))
            self._safe_set_value(ws, insert_row, 2, force.get("observation"))
            self._safe_set_value(ws, insert_row, 3, force.get("impact"))
            self._safe_set_value(ws, insert_row, 4, force.get("implications"))
            
            insert_row += 1
    
    def _map_pestel(self, ws, pestel_data: List[Dict[str, Any]]):
        """Map PESTEL data - insert rows for each observation."""
        section_row = self._find_section_start(ws, "PESTEL")
        if not section_row:
            logger.warning("PESTEL section not found")
            return
        
        data_start = self._find_data_start_row(ws, section_row)
        next_section = self._find_next_section_start(ws, section_row)
        last_row = next_section - 1 if next_section else data_start + 50
        
        insert_row = data_start
        for row_idx in range(data_start, last_row):
            if not any(ws.cell(row=row_idx, column=col).value for col in range(1, 5)):
                insert_row = row_idx
                break
            insert_row = row_idx + 1
        
        for item in pestel_data:
            if insert_row > last_row:
                self._insert_row_with_formatting(ws, insert_row, insert_row - 1)
            
            self._safe_set_value(ws, insert_row, 1, item.get("factor"))
            self._safe_set_value(ws, insert_row, 2, item.get("observation"))
            self._safe_set_value(ws, insert_row, 3, item.get("impact"))
            self._safe_set_value(ws, insert_row, 4, item.get("implications"))
            
            insert_row += 1
    
    def _map_swot(self, ws, swot_data: Dict[str, Any]):
        """Map SWOT data - one row per item in each block, Column A only.

        Items are written to Column A (the item column). Column B ("Our Actions")
        is left empty for the user to fill in. Items are truncated to fit the
        available template slots — no rows are inserted.
        """
        section_row = self._find_section_start(ws, "SWOT")
        if not section_row:
            logger.warning("SWOT section not found")
            return

        # Find each SWOT category block
        categories = ["Strengths", "Weaknesses", "Opportunities", "Threats"]
        data_keys = ["strengths", "weaknesses", "opportunities", "threats"]

        for category, data_key in zip(categories, data_keys):
            items = swot_data.get(data_key, [])
            if not items:
                continue

            # Find the category header row
            category_row = None
            for row_idx in range(section_row, section_row + 100):
                cell = ws.cell(row=row_idx, column=1)
                if cell.value and category.upper() in str(cell.value).upper():
                    category_row = row_idx
                    break

            if not category_row:
                continue

            # Data starts on the row immediately after the category header
            data_start = category_row + 1

            # Find the boundary: next category header or next major section
            next_category_row = None
            for row_idx in range(category_row + 1, category_row + 50):
                cell = ws.cell(row=row_idx, column=1)
                if cell.value and any(c.upper() in str(cell.value).upper() for c in categories if c != category):
                    next_category_row = row_idx
                    break

            # If no next category found, look for the next major section
            if not next_category_row:
                next_section = self._find_next_section_start(ws, section_row)
                next_category_row = next_section if next_section else data_start + 20

            # Available empty slots (rows between this header and the next)
            available_slots = next_category_row - data_start

            # Truncate items to fit available slots — never insert rows
            items_to_write = items[:available_slots]
            if len(items) > available_slots:
                logger.warning(
                    f"SWOT {category}: {len(items)} items but only {available_slots} "
                    f"slots available. Truncating to {available_slots}."
                )

            # Write items to Column A (item column), leave Column B for user actions
            for i, item in enumerate(items_to_write):
                self._safe_set_value(ws, data_start + i, 1, str(item))
    
    def _map_customer_analysis(self, ws, customers: List[Dict[str, Any]]):
        """Map customer analysis - one row per customer."""
        section_row = self._find_section_start(ws, "CUSTOMER ANALYSIS")
        if not section_row:
            logger.warning("CUSTOMER ANALYSIS section not found")
            return
        
        data_start = self._find_data_start_row(ws, section_row)
        next_section = self._find_next_section_start(ws, section_row)
        last_row = next_section - 1 if next_section else data_start + 50
        
        insert_row = data_start
        for row_idx in range(data_start, last_row):
            if not ws.cell(row=row_idx, column=1).value:
                insert_row = row_idx
                break
            insert_row = row_idx + 1
        
        for customer in customers:
            if insert_row > last_row:
                self._insert_row_with_formatting(ws, insert_row, insert_row - 1)
            
            self._safe_set_value(ws, insert_row, 1, customer.get("customer_name"))
            self._safe_set_value(ws, insert_row, 2, customer.get("y1_revenue"))
            self._safe_set_value(ws, insert_row, 3, customer.get("y2_revenue"))
            self._safe_set_value(ws, insert_row, 4, customer.get("y3_revenue"))
            self._safe_set_value(ws, insert_row, 5, customer.get("trend_notes"))
            self._safe_set_value(ws, insert_row, 6, customer.get("action"))
            
            insert_row += 1
    
    def _map_product_analysis(self, ws, products: List[Dict[str, Any]]):
        """Map product analysis - one row per product."""
        section_row = self._find_section_start(ws, "PRODUCT ANALYSIS")
        if not section_row:
            logger.warning("PRODUCT ANALYSIS section not found")
            return
        
        data_start = self._find_data_start_row(ws, section_row)
        next_section = self._find_next_section_start(ws, section_row)
        last_row = next_section - 1 if next_section else data_start + 50
        
        insert_row = data_start
        for row_idx in range(data_start, last_row):
            if not ws.cell(row=row_idx, column=1).value:
                insert_row = row_idx
                break
            insert_row = row_idx + 1
        
        for product in products:
            if insert_row > last_row:
                self._insert_row_with_formatting(ws, insert_row, insert_row - 1)
            
            self._safe_set_value(ws, insert_row, 1, product.get("product"))
            self._safe_set_value(ws, insert_row, 2, product.get("lifecycle_stage"))
            self._safe_set_value(ws, insert_row, 3, product.get("delivery_limitations"))
            self._safe_set_value(ws, insert_row, 4, product.get("opportunities"))
            
            insert_row += 1
    
    def _map_competitor_analysis(self, ws, competitors: List[Dict[str, Any]]):
        """Map competitor analysis - one row per competitor."""
        section_row = self._find_section_start(ws, "COMPETITOR ANALYSIS")
        if not section_row:
            logger.warning("COMPETITOR ANALYSIS section not found")
            return
        
        data_start = self._find_data_start_row(ws, section_row)
        next_section = self._find_next_section_start(ws, section_row)
        last_row = next_section - 1 if next_section else data_start + 50
        
        insert_row = data_start
        for row_idx in range(data_start, last_row):
            if not ws.cell(row=row_idx, column=1).value:
                insert_row = row_idx
                break
            insert_row = row_idx + 1
        
        for competitor in competitors:
            if insert_row > last_row:
                self._insert_row_with_formatting(ws, insert_row, insert_row - 1)
            
            self._safe_set_value(ws, insert_row, 1, competitor.get("market_segment"))
            self._safe_set_value(ws, insert_row, 2, competitor.get("competitor"))
            self._safe_set_value(ws, insert_row, 3, competitor.get("strengths"))
            self._safe_set_value(ws, insert_row, 4, competitor.get("weaknesses"))
            self._safe_set_value(ws, insert_row, 5, competitor.get("relative_size"))
            self._safe_set_value(ws, insert_row, 6, competitor.get("how_we_compete"))
            self._safe_set_value(ws, insert_row, 7, competitor.get("likely_moves"))
            
            insert_row += 1
    
    def _map_growth_opportunities(self, ws, opportunities: List[Dict[str, Any]]):
        """Map growth opportunities - insert rows with Ansoff data."""
        section_row = self._find_section_start(ws, "GROWTH OPPORTUNITIES")
        if not section_row:
            logger.warning("GROWTH OPPORTUNITIES section not found")
            return
        
        data_start = self._find_data_start_row(ws, section_row)
        next_section = self._find_next_section_start(ws, section_row)
        last_row = next_section - 1 if next_section else data_start + 50
        
        insert_row = data_start
        for row_idx in range(data_start, last_row):
            if not ws.cell(row=row_idx, column=1).value:
                insert_row = row_idx
                break
            insert_row = row_idx + 1
        
        for opp in opportunities:
            if insert_row > last_row:
                self._insert_row_with_formatting(ws, insert_row, insert_row - 1)
            
            self._safe_set_value(ws, insert_row, 1, opp.get("category"))
            self._safe_set_value(ws, insert_row, 2, opp.get("segment"))
            self._safe_set_value(ws, insert_row, 3, opp.get("action"))
            self._safe_set_value(ws, insert_row, 4, opp.get("time_horizon"))
            self._safe_set_value(ws, insert_row, 5, opp.get("success_chance"))
            self._safe_set_value(ws, insert_row, 6, opp.get("notes"))
            
            insert_row += 1
    
    def _map_financial_targets(self, ws, financial_data: Dict[str, Any]):
        """Map financial targets to specific rows."""
        section_row = self._find_section_start(ws, "FINANCIAL TARGET")
        if not section_row:
            logger.warning("FINANCIAL TARGET section not found")
            return
        
        # Find rows for current FY and next FY
        current_fy = financial_data.get("current_fy", {})
        next_fy = financial_data.get("next_fy", {})
        
        # Search for "Current FY" and "Next FY" labels
        for row_idx in range(section_row, section_row + 20):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value:
                text = str(cell.value).upper()
                if "CURRENT" in text and "FY" in text:
                    # Map revenue, GP, NP to columns
                    if current_fy.get("revenue"):
                        self._safe_set_value(ws, row_idx, 2, current_fy.get("revenue"))
                    if current_fy.get("gross_profit"):
                        self._safe_set_value(ws, row_idx, 3, current_fy.get("gross_profit"))
                    if current_fy.get("net_profit"):
                        self._safe_set_value(ws, row_idx, 4, current_fy.get("net_profit"))
                elif "NEXT" in text and "FY" in text:
                    if next_fy.get("revenue"):
                        self._safe_set_value(ws, row_idx, 2, next_fy.get("revenue"))
                    if next_fy.get("gross_profit"):
                        self._safe_set_value(ws, row_idx, 3, next_fy.get("gross_profit"))
                    if next_fy.get("net_profit"):
                        self._safe_set_value(ws, row_idx, 4, next_fy.get("net_profit"))
    
    def _map_risks(self, ws, risks_data: Dict[str, Any]):
        """Map risks to category rows in Column B."""
        section_row = self._find_section_start(ws, "RISKS")
        if not section_row:
            logger.warning("RISKS section not found")
            return
        
        categories = ["Legal", "Financial", "Operations", "People", "SM", "Product", "Other"]
        data_keys = ["legal", "financial", "operations", "people", "sm", "product", "other"]
        
        for category, data_key in zip(categories, data_keys):
            risk_items = risks_data.get(data_key, [])
            if not risk_items:
                continue
            
            # Find the category row
            for row_idx in range(section_row, section_row + 50):
                cell = ws.cell(row=row_idx, column=1)
                if cell.value and category.upper() in str(cell.value).upper():
                    # Write risks to Column B, using line breaks
                    existing = ws.cell(row=row_idx, column=2).value or ""
                    new_risks = "\n".join(str(r) for r in risk_items if r)
                    combined = existing + "\n" + new_risks if existing else new_risks
                    self._safe_set_value(ws, row_idx, 2, combined)
                    break
    
    def _map_strategic_priorities(self, ws, priorities: List[Dict[str, Any]]):
        """Map strategic priorities - insert rows with all columns."""
        section_row = self._find_section_start(ws, "STRATEGIC PRIORITIES")
        if not section_row:
            logger.warning("STRATEGIC PRIORITIES section not found")
            return
        
        data_start = self._find_data_start_row(ws, section_row)
        next_section = self._find_next_section_start(ws, section_row)
        last_row = next_section - 1 if next_section else data_start + 50
        
        insert_row = data_start
        for row_idx in range(data_start, last_row):
            if not ws.cell(row=row_idx, column=1).value:
                insert_row = row_idx
                break
            insert_row = row_idx + 1
        
        for priority in priorities:
            if insert_row > last_row:
                self._insert_row_with_formatting(ws, insert_row, insert_row - 1)
            
            self._safe_set_value(ws, insert_row, 1, priority.get("priority_theme"))
            self._safe_set_value(ws, insert_row, 2, priority.get("objective"))
            self._safe_set_value(ws, insert_row, 3, priority.get("initiative"))
            self._safe_set_value(ws, insert_row, 4, priority.get("owner"))
            self._safe_set_value(ws, insert_row, 5, priority.get("timeframe"))
            self._safe_set_value(ws, insert_row, 6, priority.get("kpi"))
            
            insert_row += 1
    
    def _map_key_actions(self, ws, actions: List[Dict[str, Any]]):
        """Map key actions - insert rows with action items."""
        section_row = self._find_section_start(ws, "KEY ACTIONS")
        if not section_row:
            logger.warning("KEY ACTIONS section not found")
            return
        
        data_start = self._find_data_start_row(ws, section_row)
        next_section = self._find_next_section_start(ws, section_row)
        last_row = next_section - 1 if next_section else data_start + 50
        
        insert_row = data_start
        for row_idx in range(data_start, last_row):
            if not ws.cell(row=row_idx, column=1).value:
                insert_row = row_idx
                break
            insert_row = row_idx + 1
        
        for action in actions:
            if insert_row > last_row:
                self._insert_row_with_formatting(ws, insert_row, insert_row - 1)
            
            self._safe_set_value(ws, insert_row, 1, action.get("action_item"))
            self._safe_set_value(ws, insert_row, 2, action.get("notes"))
            
            insert_row += 1


def get_strategy_workbook_exporter(template_path: Optional[Path] = None) -> StrategyWorkbookExporter:
    """Factory/helper for dependency injection."""
    return StrategyWorkbookExporter(template_path)

