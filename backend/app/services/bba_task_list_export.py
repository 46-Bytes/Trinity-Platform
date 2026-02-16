"""
Phase 2 – Excel Task List (Engagement Planner) exporter.

Given task rows and summary data from the BBA task planner service, this module
creates an Excel workbook with:
- Main sheet of tasks with columns:
  Rec #, Recommendation, Owner, Task, Advisor Hrs, Advisor, Status, Notes, Timing
- Data validation for Status (dropdown)
- Readable formatting (column widths, header styling, alternating row shading)
- Optional summary sheet with monthly BBA hours
- Footer: "Prepared by Benchmark Business Advisory – Confidential."
"""

from __future__ import annotations

from typing import List, Dict, Any
import io
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - dependency may not be installed in all environments
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class BBATaskListExporter:
    """
    Create an Excel (.xlsx) advisor task list from task planner rows.
    """

    STATUS_OPTIONS = [
        "Not yet started",
        "In progress",
        "Complete",
        "Awaiting review",
    ]

    OWNER_OPTIONS = ["Client", "BBA"]

    def __init__(self) -> None:
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #

    def generate_workbook_bytes(
        self,
        tasks: List[Dict[str, Any]],
        summary: Dict[str, Any] | None = None,
    ) -> bytes:
        """
        Build an Excel workbook for the given tasks and optional summary.

        Args:
            tasks: List of task dictionaries compatible with BBATaskRow
            summary: Optional summary dictionary from BBATaskPlannerService

        Returns:
            Excel file bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Advisor Task List"

        self._write_header_row(ws)
        self._write_task_rows(ws, tasks)
        last_row = ws.max_row
        self._apply_owner_validation(ws, first_row=2, last_row=last_row)
        self._apply_advisor_validation(ws, first_row=2, last_row=last_row, tasks=tasks)
        self._apply_status_validation(ws, first_row=2, last_row=last_row)
        self._apply_basic_formatting(ws)
        self._apply_status_conditional_formatting(ws, first_row=2, last_row=ws.max_row)
        self._set_footer(ws)

        if summary:
            self._add_summary_sheet(wb, summary)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # ------------------------------------------------------------------ #
    # Internal helpers – main sheet
    # ------------------------------------------------------------------ #

    def _write_header_row(self, ws) -> None:
        headers = [
            "Rec #",
            "Recommendation",
            "Owner",
            "Task",
            "Advisor Hrs",
            "Advisor",
            "Status",
            "Notes",
            "Timing",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1a365d")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

    def _write_task_rows(self, ws, tasks: List[Dict[str, Any]]) -> None:
        for task in tasks:
            ws.append(
                [
                    task.get("rec_number"),
                    task.get("recommendation"),
                    task.get("owner"),
                    task.get("task"),
                    task.get("advisorHrs"),
                    task.get("advisor"),
                    task.get("status") or self.STATUS_OPTIONS[0],
                    task.get("notes") or "",
                    task.get("timing"),
                ]
            )

    def _apply_owner_validation(self, ws, first_row: int, last_row: int) -> None:
        """Add Excel data validation (dropdown) to the Owner column (C)."""
        if last_row < first_row:
            return
        # Inline list so dropdown arrow shows reliably in Excel
        options_str = ",".join(self.OWNER_OPTIONS)
        formula1 = f'"{options_str}"'
        dv = DataValidation(
            type="list",
            formula1=formula1,
            allow_blank=False,
            showDropDown=True,
            showInputMessage=True,
            promptTitle="Owner",
            prompt="Select Client or BBA",
        )
        ws.add_data_validation(dv)
        dv.add(f"C{first_row}:C{last_row}")

    def _apply_advisor_validation(
        self, ws, first_row: int, last_row: int, tasks: List[Dict[str, Any]]
    ) -> None:
        """Add Excel data validation (dropdown) to the Advisor column (F) from task names."""
        if last_row < first_row:
            return
        advisor_names = sorted(
            {str(t.get("advisor") or "").strip() for t in tasks if t.get("advisor")}
        )
        if not advisor_names:
            return
        options_str = ",".join(advisor_names)
        formula1 = f'"{options_str}"'
        dv = DataValidation(
            type="list",
            formula1=formula1,
            allow_blank=True,
            showDropDown=True,
            showInputMessage=True,
            promptTitle="Advisor",
            prompt="Select an advisor",
        )
        ws.add_data_validation(dv)
        dv.add(f"F{first_row}:F{last_row}")

    def _apply_status_validation(self, ws, first_row: int, last_row: int) -> None:
        """
        Add Excel data validation (dropdown) to the Status column (G) with
        options: Not yet started, In progress, Complete, Awaiting review.
        Uses inline list so the dropdown arrow shows when the cell is selected.
        """
        if last_row < first_row:
            return
        # Inline list: comma-separated values in quotes (Excel list validation)
        options_str = ",".join(self.STATUS_OPTIONS)
        formula1 = f'"{options_str}"'
        dv = DataValidation(
            type="list",
            formula1=formula1,
            allow_blank=False,
            showDropDown=True,
            showInputMessage=True,
            promptTitle="Status",
            prompt="Select a status from the list",
        )
        dv.error = "Please choose a status from the dropdown list."
        dv.errorTitle = "Invalid Status"
        ws.add_data_validation(dv)
        dv.add(f"G{first_row}:G{last_row}")

    def _apply_basic_formatting(self, ws) -> None:
        # Column widths (approximate for readability)
        widths = {
            "A": 6,   # Rec #
            "B": 32,  # Recommendation
            "C": 12,  # Owner
            "D": 60,  # Task
            "E": 12,  # Advisor Hrs
            "F": 20,  # Advisor
            "G": 18,  # Status
            "H": 26,  # Notes
            "I": 20,  # Timing
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        # Alternating row shading for readability
        fill_alt = PatternFill("solid", fgColor="f7fafc")
        # Status (G) and Timing (I): same "dropdown" style - light fill so they match in the xlsx
        fill_dropdown = PatternFill("solid", fgColor="E8F4F8")
        status_col = 7   # G
        timing_col = 9   # I
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = fill_alt
            # Apply same dropdown-style fill to Status and Timing
            ws.cell(row=row, column=status_col).fill = fill_dropdown
            ws.cell(row=row, column=timing_col).fill = fill_dropdown

        # Wrap text for Task, Notes, Recommendation
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        for row in range(2, ws.max_row + 1):
            for col in (2, 4, 8):  # B, D, H
                ws.cell(row=row, column=col).alignment = wrap_alignment

        # Number format for Advisor Hrs
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=5).number_format = "0.0##"

        # Thin borders around the table
        thin = Side(border_style="thin", color="DDDDDD")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in range(1, ws.max_row + 1):
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border

    def _apply_status_conditional_formatting(
        self,
        ws,
        first_row: int,
        last_row: int,
    ) -> None:
        """
        Optional: colour‑code status values.
        Uses simple text comparisons against the Status column.
        """
        if last_row < first_row:
            return

        status_col = "G"
        cell_range = f"{status_col}{first_row}:{status_col}{last_row}"

        # Green for Complete
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"Complete"'],
                stopIfTrue=False,
                fill=PatternFill("solid", fgColor="c6f6d5"),
            ),
        )
        # Yellow for In progress
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"In progress"'],
                stopIfTrue=False,
                fill=PatternFill("solid", fgColor="faf089"),
            ),
        )
        # Blue for Awaiting review
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"Awaiting review"'],
                stopIfTrue=False,
                fill=PatternFill("solid", fgColor="bee3f8"),
            ),
        )

    def _set_footer(self, ws) -> None:
        """
        Add confidential footer note.
        """
        try:
            ws.oddFooter.center.text = (
                "Prepared by Benchmark Business Advisory – Confidential."
            )
        except Exception:
            # Footer is non‑critical; ignore if worksheet footer is unavailable
            logger.debug("Could not set Excel footer for task list sheet", exc_info=True)

    # ------------------------------------------------------------------ #
    # Optional summary sheet
    # ------------------------------------------------------------------ #

    def _add_summary_sheet(self, wb: "Workbook", summary: Dict[str, Any]) -> None:
        """
        Add a simple summary sheet with monthly BBA hours and any warnings.
        """
        ws = wb.create_sheet(title="Summary")

        ws["A1"] = "Monthly BBA Hours"
        ws["A1"].font = Font(bold=True, size=14)

        ws.append(["Month (YYYY-MM)", "Total BBA Hours", "Capacity (Hours)"])
        header_font = Font(bold=True)
        for col in range(1, 4):
            ws.cell(row=2, column=col).font = header_font

        monthly_hours = summary.get("monthly_hours") or {}
        capacity = summary.get("max_hours_per_month")

        row_idx = 3
        for ym, hours in sorted(monthly_hours.items()):
            ws.cell(row=row_idx, column=1).value = ym
            ws.cell(row=row_idx, column=2).value = float(hours)
            ws.cell(row=row_idx, column=3).value = float(capacity)
            row_idx += 1

        # Autosize first three columns
        for col_letter in ("A", "B", "C"):
            ws.column_dimensions[col_letter].width = 20

        # Warnings section
        warnings = summary.get("warnings") or []
        if warnings:
            start_row = row_idx + 2
            ws.cell(row=start_row, column=1).value = "Warnings"
            ws.cell(row=start_row, column=1).font = Font(bold=True)

            for offset, warning in enumerate(warnings, start=1):
                ws.cell(row=start_row + offset, column=1).value = f"- {warning}"
                ws.cell(start_row + offset, column=1).alignment = Alignment(
                    wrap_text=True
                )


def get_bba_task_list_exporter() -> BBATaskListExporter:
    """Factory/helper for dependency injection."""
    return BBATaskListExporter()

