"""
Half-Yearly Role Scorecard Excel exporter.

Writes the strict single-sheet template: Role Purpose, then Responsibilities &
Outcomes, Behaviour & Leadership Expectations, Transition Milestones and the
Half-Yearly Summary.

Rating dropdowns are applied by section and column letter, using the row ranges
this exporter itself wrote. Header text is never matched, so a dropdown cannot
land on a comments column.
"""
from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple
import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

SHEET_NAME = "H1-H2 Scorecard"

# The widest section is Responsibilities & Outcomes at nine columns (A-I).
LAST_COLUMN = 9

RATING_OPTIONS = '"1,2,3,4,5"'

# Section 1 is nine columns; sections 2 and 3 are eight and therefore sit one
# column to the left from the ratings onward.
RESPONSIBILITY_HEADINGS = [
    "Focus Area",
    "Core Accountability",
    "Performance Indicators / Outcomes",
    "H1 Self (1-5)",
    "H1 Mgr (1-5)",
    "H1 Comments",
    "H2 Self (1-5)",
    "H2 Mgr (1-5)",
    "H2 Comments",
]
RESPONSIBILITY_RATING_COLUMNS = ["D", "E", "G", "H"]

BEHAVIOUR_HEADINGS = [
    "Behavioural Focus",
    "Expected Demonstration",
    "H1 Self (1-5)",
    "H1 Mgr (1-5)",
    "H1 Comments",
    "H2 Self (1-5)",
    "H2 Mgr (1-5)",
    "H2 Comments",
]
MILESTONE_HEADINGS = [
    "Milestone",
    "Target Date",
    "H1 Self (1-5)",
    "H1 Mgr (1-5)",
    "H1 Comments",
    "H2 Self (1-5)",
    "H2 Mgr (1-5)",
    "H2 Comments",
]
# Shared by sections 2 and 3, which have the same eight-column shape.
EIGHT_COLUMN_RATING_COLUMNS = ["C", "D", "F", "G"]

# Minimum widths. A column can be free text in one section and a rating in
# another, so each takes the widest role it plays: C is text in section 1,
# E, F and H hold comments in one section or another.
MIN_COLUMN_WIDTHS = {
    "A": 28, "B": 40, "C": 40, "D": 10, "E": 22,
    "F": 22, "G": 10, "H": 22, "I": 22,
}
MAX_COLUMN_WIDTH = 60
WIDTH_PADDING = 2

# Free-text columns wrap rather than forcing an unusably wide sheet.
WRAP_COLUMNS = {"A", "B", "C"}


class ScorecardExporter:
    """Generates the Half-Yearly Role Scorecard workbook for a single role."""

    def generate_workbook_bytes(
        self,
        role_title: str,
        scorecard_content: Dict[str, Any],
        fy_range: Optional[str] = None,
    ) -> io.BytesIO:
        """
        Build the .xlsx for one role.

        Args:
            role_title: The role the scorecard covers, used for logging only —
                the strict template carries no title row.
            scorecard_content: The approved scorecard sections.
            fy_range: Financial year range shown in the Transition Milestones heading.

        Returns:
            BytesIO stream positioned at the start, ready to stream to the client.
        """
        content = scorecard_content or {}

        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        row = self._write_role_purpose(ws, content.get("role_purpose"))
        row += 1  # blank row

        row = self._write_responsibilities(ws, content.get("responsibilities") or [], row)
        row = self._write_behaviours(ws, content.get("behaviours") or [], row)
        row = self._write_milestones(ws, content.get("milestones") or [], row, fy_range)
        self._write_summary(ws, row)

        self._fit_columns(ws)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        logger.info(f"Generated scorecard workbook for '{role_title}'")
        return stream

    # ==================== Sections ====================

    def _write_role_purpose(self, ws, role_purpose: Any) -> int:
        """
        Rows 1-2: the label and the purpose text. Row 2 is the only row in the
        sheet permitted to span columns.
        """
        label = ws.cell(row=1, column=1)
        label.value = "Role Purpose"
        label.font = Font(bold=True)
        label.alignment = Alignment(horizontal="left")

        text = ws.cell(row=2, column=1)
        text.value = self._clean(role_purpose)
        text.alignment = Alignment(vertical="top", wrap_text=True)

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=LAST_COLUMN)

        return 3

    def _write_responsibilities(self, ws, rows: List[Any], start_row: int) -> int:
        """Section 1: nine columns, ratings on D, E, G and H."""
        if not rows:
            return start_row

        row = self._write_section_header(
            ws, start_row, "Responsibilities and Outcomes (Half-Yearly Review - Self vs Manager)"
        )
        row = self._write_column_headings(ws, row, RESPONSIBILITY_HEADINGS)

        first_data_row = row
        for entry in rows:
            if not isinstance(entry, dict):
                continue
            self._write_cell(ws, row, 1, entry.get("focus_area"))
            self._write_cell(ws, row, 2, entry.get("core_accountability"))
            self._write_cell(ws, row, 3, entry.get("performance_indicators"))
            row += 1

        self._apply_rating_validation(ws, RESPONSIBILITY_RATING_COLUMNS, first_data_row, row - 1)

        return row + 1  # trailing blank row

    def _write_behaviours(self, ws, rows: List[Any], start_row: int) -> int:
        """Section 2: eight columns, ratings on C, D, F and G."""
        if not rows:
            return start_row

        row = self._write_section_header(ws, start_row, "Behaviour and Leadership Expectations")
        row = self._write_column_headings(ws, row, BEHAVIOUR_HEADINGS)

        first_data_row = row
        for entry in rows:
            if not isinstance(entry, dict):
                continue
            self._write_cell(ws, row, 1, entry.get("behavioural_focus"))
            self._write_cell(ws, row, 2, entry.get("expected_demonstration"))
            row += 1

        self._apply_rating_validation(ws, EIGHT_COLUMN_RATING_COLUMNS, first_data_row, row - 1)

        return row + 1

    def _write_milestones(
        self, ws, rows: List[Any], start_row: int, fy_range: Optional[str]
    ) -> int:
        """Section 3: eight columns, ratings on C, D, F and G. Omitted when empty."""
        if not rows:
            return start_row

        fy_label = f" ({fy_range} - Self vs Manager)" if fy_range else " (Self vs Manager)"
        row = self._write_section_header(ws, start_row, f"Transition Milestones{fy_label}")
        row = self._write_column_headings(ws, row, MILESTONE_HEADINGS)

        first_data_row = row
        for entry in rows:
            if not isinstance(entry, dict):
                continue
            self._write_cell(ws, row, 1, entry.get("milestone"))
            self._write_cell(ws, row, 2, entry.get("target_date"))
            row += 1

        self._apply_rating_validation(ws, EIGHT_COLUMN_RATING_COLUMNS, first_data_row, row - 1)

        return row + 1

    def _write_summary(self, ws, start_row: int) -> None:
        """Section 4: column A only, every other cell left blank."""
        row = self._write_section_header(ws, start_row, "Half-Yearly Summary")

        for label in (
            "Average Self Rating:",
            "Average Manager Rating:",
            "Key Discussion Points / Agreed Actions:",
        ):
            ws.cell(row=row, column=1).value = label
            row += 1

    # ==================== Building blocks ====================

    @staticmethod
    def _write_section_header(ws, row: int, text: str) -> int:
        """Bold, left-aligned section header on its own row."""
        cell = ws.cell(row=row, column=1)
        cell.value = text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
        return row + 1

    @staticmethod
    def _write_column_headings(ws, row: int, headings: List[str]) -> int:
        """Write a section's column headings starting at column A."""
        for offset, heading in enumerate(headings, start=1):
            cell = ws.cell(row=row, column=offset)
            cell.value = heading
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        return row + 1

    def _write_cell(self, ws, row: int, column: int, value: Any) -> None:
        """Write a text cell, wrapping the free-text columns."""
        cell = ws.cell(row=row, column=column)
        cell.value = self._clean(value)
        if get_column_letter(column) in WRAP_COLUMNS:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    @staticmethod
    def _apply_rating_validation(
        ws, columns: List[str], first_row: int, last_row: int
    ) -> None:
        """
        Attach a 1-5 dropdown to the given columns across a section's data rows.

        Columns are passed in by the caller that wrote the section, so validation
        follows the section's shape rather than any header text.
        """
        if last_row < first_row:
            return

        validation = DataValidation(
            type="list",
            formula1=RATING_OPTIONS,
            allow_blank=True,
            showDropDown=True,
        )
        ws.add_data_validation(validation)

        for column in columns:
            validation.add(f"{column}{first_row}:{column}{last_row}")

    @staticmethod
    def _fit_columns(ws) -> None:
        """
        Widen columns to fit their content, never below the per-column minimum
        and never past MAX_COLUMN_WIDTH. Wrapping handles anything longer.
        """
        for column in range(1, LAST_COLUMN + 1):
            letter = get_column_letter(column)

            longest = 0
            for row in range(1, ws.max_row + 1):
                # Merged cells report their value on the anchor only; row 2 is
                # the sole merge and is excluded so it cannot skew the width.
                if row == 2:
                    continue
                value = ws.cell(row=row, column=column).value
                if value:
                    longest = max(longest, len(str(value)))

            needed = longest + WIDTH_PADDING if longest else 0
            minimum = MIN_COLUMN_WIDTHS.get(letter, 12)
            ws.column_dimensions[letter].width = min(max(minimum, needed), MAX_COLUMN_WIDTH)

    @staticmethod
    def _clean(value: Any) -> Optional[str]:
        """
        Normalise a cell value. Blanks stay blank — nothing is substituted for
        missing information.
        """
        if value is None:
            return None
        text = str(value).strip()
        return text or None


_exporter: Optional[ScorecardExporter] = None


def get_scorecard_exporter() -> ScorecardExporter:
    """Get the shared scorecard exporter instance"""
    global _exporter
    if _exporter is None:
        _exporter = ScorecardExporter()
    return _exporter
