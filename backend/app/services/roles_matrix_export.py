"""
Roles & Responsibilities Matrix Excel exporter.

Writes the reviewed matrix rows into a copy of the HR Planning Tool workbook so
the output keeps the exact "Job Roles" layout, headings and formatting.
"""
from __future__ import annotations

from typing import Any, Dict, List, Optional
from copy import copy
from pathlib import Path
import io
import logging

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

SHEET_NAME = "Job Roles"

# Layout of the Job Roles sheet:
#   row 1  - "HR MANAGEMENT PLAN" banner (A1:J1 merged)
#   row 2  - "Role Analysis" banner (A2:J2 merged)
#   row 3  - column headings
#   row 4+ - data
FIRST_DATA_ROW = 4
LAST_COLUMN = 10  # J

# A plain data row inside the styled region, used as the formatting source when
# the matrix runs past the rows the template already styles.
STYLE_SOURCE_ROW = 5

# Row keys in column order A-J
ROW_KEYS = [
    "name",
    "role_description",
    "time",
    "priorities",
    "retain",
    "gain",
    "lose",
    "action",
    "resp",
    "when",
]


class RolesMatrixExporter:
    """Generates the completed HR Planning Tool workbook from matrix rows."""

    def __init__(self, template_path: Optional[Path] = None):
        """
        Args:
            template_path: Path to the template file. Defaults to the copy stored
                in backend/files/templates/roles-matrix/.
        """
        if template_path is None:
            base_dir = Path(__file__).resolve().parents[2]  # backend/
            template_path = base_dir / "files" / "templates" / "roles-matrix" / "HR Planning Tool.xlsx"

        self.template_path = Path(template_path)

        if not self.template_path.exists():
            raise FileNotFoundError(
                f"Template file not found at {self.template_path}. "
                f"Please place 'HR Planning Tool.xlsx' in the templates directory."
            )

    def generate_workbook_bytes(self, matrix_rows: List[Dict[str, Any]]) -> io.BytesIO:
        """
        Build the workbook from the given matrix rows.

        Args:
            matrix_rows: Ordered rows, each keyed by the Job Roles column names.

        Returns:
            BytesIO stream positioned at the start, ready to stream to the client.
        """
        wb = load_workbook(self.template_path)

        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Template is missing the '{SHEET_NAME}' sheet")

        ws = wb[SHEET_NAME]

        self._unmerge_data_region(ws)
        last_styled_row = self._find_last_styled_row(ws)
        self._clear_data_region(ws)
        self._write_rows(ws, matrix_rows, last_styled_row)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        logger.info(f"Generated roles matrix workbook with {len(matrix_rows)} row(s)")
        return stream

    # ==================== Template preparation ====================

    @staticmethod
    def _unmerge_data_region(ws) -> None:
        """
        Drop merged ranges inside the data region.

        The sample workbook merges the Time column across one person's block
        (C38:C43). Rows are written one per line, so those merges are removed.
        The two banner merges on rows 1-2 are left intact.
        """
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row >= FIRST_DATA_ROW:
                ws.unmerge_cells(str(merged_range))

    @staticmethod
    def _find_last_styled_row(ws) -> int:
        """
        Find the last row the template has already formatted, by walking down the
        Role Descriptions column until the cell borders stop.
        """
        last_styled = FIRST_DATA_ROW
        for row in range(FIRST_DATA_ROW, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)
            if cell.border and cell.border.bottom and cell.border.bottom.style:
                last_styled = row
            else:
                break
        return last_styled

    @staticmethod
    def _clear_data_region(ws) -> None:
        """Remove the sample data, leaving the formatting in place."""
        for row in range(FIRST_DATA_ROW, ws.max_row + 1):
            for column in range(1, LAST_COLUMN + 1):
                ws.cell(row=row, column=column).value = None

    # ==================== Writing ====================

    def _write_rows(self, ws, matrix_rows: List[Dict[str, Any]], last_styled_row: int) -> None:
        """Write one worksheet row per matrix row, from row 4 down."""
        for offset, row_data in enumerate(matrix_rows or []):
            target_row = FIRST_DATA_ROW + offset

            if target_row > last_styled_row:
                self._copy_row_style(ws, source_row=STYLE_SOURCE_ROW, target_row=target_row)

            for column, key in enumerate(ROW_KEYS, start=1):
                value = (row_data or {}).get(key)
                ws.cell(row=target_row, column=column).value = self._clean(value)

    @staticmethod
    def _copy_row_style(ws, source_row: int, target_row: int) -> None:
        """Carry the template's data-row formatting onto a row beyond the styled region."""
        for column in range(1, LAST_COLUMN + 1):
            source_cell = ws.cell(row=source_row, column=column)
            target_cell = ws.cell(row=target_row, column=column)
            target_cell._style = copy(source_cell._style)

        source_height = ws.row_dimensions[source_row].height
        if source_height:
            ws.row_dimensions[target_row].height = source_height

    @staticmethod
    def _clean(value: Any) -> Optional[str]:
        """
        Normalise a cell value. Blanks stay blank — nothing is substituted for
        missing information.
        """
        if value is None:
            return None
        text = str(value).strip()
        return text or None


_exporter: Optional[RolesMatrixExporter] = None


def get_roles_matrix_exporter() -> RolesMatrixExporter:
    """Get the shared roles matrix exporter instance"""
    global _exporter
    if _exporter is None:
        _exporter = RolesMatrixExporter()
    return _exporter
